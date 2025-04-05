from django.shortcuts import render
from django.views.generic import ListView, CreateView
from django.http import HttpResponse
from openpyxl import Workbook
from library.models import Author, Book, BorrowRecord
from library.forms import AuthorForm, BookForm, BorrowRecordForm

def home_view(request):
    return render(request, 'library/home.html')

class AuthorListView(ListView):
    model = Author
    template_name = 'library/author_list.html'
    context_object_name = 'authors'
    paginate_by = 10
    ordering = ['name']  # Fixes pagination warning

class BookListView(ListView):
    model = Book
    template_name = 'library/book_list.html'
    context_object_name = 'books'
    paginate_by = 10
    ordering = ['title']  # Fixes pagination warning

class BorrowRecordListView(ListView):
    model = BorrowRecord
    template_name = 'library/borrowrecord_list.html'
    context_object_name = 'borrow_records'
    paginate_by = 10
    ordering = ['-borrow_date']  # Fixes pagination warning

class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'library/forms/author_form.html'
    success_url = '/authors/'

class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'library/forms/book_form.html'
    success_url = '/books/'

class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'library/forms/borrowrecord_form.html'
    success_url = '/borrow-records/'

def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="library_data.xlsx"'
    
    wb = Workbook()
    
    # Authors sheet
    ws_authors = wb.active
    ws_authors.title = "Authors"
    ws_authors.append(['ID', 'Name', 'Email', 'Bio'])
    for author in Author.objects.all().order_by('name'):
        ws_authors.append([author.id, author.name, author.email, author.bio])
    
    # Books sheet
    ws_books = wb.create_sheet(title="Books")
    ws_books.append(['ID', 'Title', 'Genre', 'Published Date', 'Author'])
    for book in Book.objects.all().order_by('title'):
        ws_books.append([book.id, book.title, book.genre, book.published_date, str(book.author)])
    
    # Borrow Records sheet
    ws_borrow = wb.create_sheet(title="Borrow Records")
    ws_borrow.append(['ID', 'User Name', 'Book', 'Borrow Date', 'Return Date'])
    for record in BorrowRecord.objects.all().order_by('-borrow_date'):
        ws_borrow.append([record.id, record.user_name, str(record.book), record.borrow_date, record.return_date])
    
    wb.save(response)
    return response